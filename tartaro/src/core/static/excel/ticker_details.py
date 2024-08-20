from typing import Any

from openpyxl.worksheet.worksheet import Worksheet


class TickerDetailsExcelFile:
    def __init__(self, sheet: Worksheet, main_sheet: Worksheet, main_sheet_row: int):
        self.parse_methods = {
            dict: self.parse_dict,
            list: self.parse_list,
        }
        self.main_sheet = main_sheet
        self.main_sheet_row = main_sheet_row
        self.sheet = sheet
        self.row = 1

    async def parse_dict(self, value: dict, column: int):
        self.row += 1
        self.main_sheet_row += 1
        await self.populate_sheet(value, column)

    async def parse_list(self, value: list, column: int):
        for index, item in enumerate(value):
            method = self.parse_methods.get(type(item), self.set_sheet_cell)
            await method(item, column+index)

    async def set_sheet_cell(self, value: Any, column: int):
        self.sheet.cell(row=self.row, column=column).value = value
        self.main_sheet.cell(row=self.main_sheet_row, column=column).value = value

    async def populate_sheet(self, values: Any, column: int = 1):
        for field, value in values.items():
            self.sheet.cell(row=self.row, column=column).value = field
            self.main_sheet.cell(row=self.main_sheet_row, column=column).value = field
            method = self.parse_methods.get(type(value), self.set_sheet_cell)
            await method(value, column+1)
            self.row += 1
            self.main_sheet_row += 1
        return self.main_sheet_row

