from io import BytesIO

import openpyxl

from src.core.dto.graph import GraphSample
from src.core.static.excel.graph import GraphExcelFile
from src.core.static.excel.ticker_details import TickerDetailsExcelFile
from src.repository.abstract.cache import AbstractCacheRepository, async_cached_operation


class ExcelTransformation(AbstractCacheRepository):

    def __init__(self, redis_infrastructure):
        self.cache_connection = redis_infrastructure

    @async_cached_operation(ttl_in_seconds=25)
    async def get_ticker_excel_file(self, ticker_fields: dict) -> bytes:
        file = openpyxl.Workbook()
        file.remove(file.get_sheet_by_name("Sheet"))
        main_sheet = file.create_sheet("Complete")
        main_sheet_row = 1
        for collection, ticker_fields in ticker_fields.items():
            main_sheet.cell(row=main_sheet_row, column=1).value = collection
            main_sheet_row += 2
            sheet = file.create_sheet(collection)
            collection_sheet = TickerDetailsExcelFile(sheet, main_sheet, main_sheet_row)
            main_sheet_row = await collection_sheet.populate_sheet(ticker_fields) + 3
        ticker_excel = BytesIO()
        file.save(ticker_excel)
        return ticker_excel.getvalue()

    # @async_cached_operation(ttl_in_seconds=120)
    async def get_graph_excel_file(self, graph_name: str, graph_fields: GraphSample) -> bytes:
        if isinstance(graph_fields, dict):
            graph_fields = GraphSample(**graph_fields)
        values = graph_fields.sample
        columns = graph_fields.possible_fields
        enumerated_columns = {index+1: field for index, field in enumerate(columns)}
        file = openpyxl.Workbook()
        file.remove(file.get_sheet_by_name("Sheet"))
        graph_sheet = file.create_sheet(graph_name)
        await GraphExcelFile.populate_sheet(graph_sheet, values, enumerated_columns)
        ticker_excel = BytesIO()
        file.save(ticker_excel)
        return ticker_excel.getvalue()
